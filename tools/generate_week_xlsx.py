#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""週次シナリオxlsx(統合テンプレート)を現行データから生成する。
使い方: python3 tools/generate_week_xlsx.py <週開始日 YYYY-MM-DD> <出力先フォルダ>
タブ構成: meta / route / sections(音声区間) / events / placements
"""
import sys, json, re, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

def load_js_array(path, name):
    s = open(path, encoding='utf-8').read()
    m = re.search(name + r'\s*=\s*(\[.*?\]);', s, re.S)
    return json.loads(m.group(1))

def main(week, outdir):
    route = load_js_array('route-data.js', 'const ROUTE_POINTS')
    events = load_js_array('route-data.js', 'const EVENT_POINTS')
    narr  = load_js_array('narration.js', 'const NARRATION')

    wb = Workbook()
    head_font = Font(bold=True, color='7A4A1D')
    head_fill = PatternFill('solid', fgColor='FDF1D7')
    def sheet(name, headers, rows):
        ws = wb.create_sheet(name)
        ws.append(headers)
        for c in ws[1]:
            c.font = head_font; c.fill = head_fill
        for r in rows: ws.append(r)
        for col in ws.columns:
            w = max(len(str(c.value)) if c.value else 0 for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max(w + 2, 10), 60)
        return ws

    wb.remove(wb.active)
    sheet('meta', ['key', 'value', '説明'], [
        ['week_start', week, '週の開始日(日曜)。ファイル名と一致させる'],
        ['title', '箱庭！バスラジオ 歴史ウィーク', '週のタイトル'],
        ['theme', 'evening', 'evening / morning / night'],
        ['loop_minutes', 30, '1周の時間(分)'],
        ['start_voice', 'voice.wav', '運行開始時の音声(任意)'],
        ['bgm', '', 'ループBGM(任意)'],
    ])
    sheet('route', ['no', 'name', 'category', 'lat', 'lng'],
          [[p.get('no'), p.get('name'), p['category'], p['lat'], p['lng']] for p in route])
    # 音声区間: 現行ナレーション点を開始点として変換(終了点=次の開始点、最後は空欄)
    rows = []
    for i, n in enumerate(narr):
        nxt = narr[i + 1] if i + 1 < len(narr) else {}
        rows.append([i + 1, f'区間{i+1}', n.get('lat'), n.get('lng'),
                     nxt.get('lat', ''), nxt.get('lng', ''), '', n.get('text', '')])
    sheet('sections', ['no', 'name', 'start_lat', 'start_lng', 'end_lat', 'end_lng', 'audio_file', 'text'], rows)
    sheet('events', ['no', 'trigger_lat', 'trigger_lng', 'radius', 'message', 'url', 'icon',
                     'target_lat', 'target_lng', 'route_points'],
          [[i + 1, e['trigger']['lat'], e['trigger']['lng'], 30, e.get('message', ''), '',
            (e.get('eventPoint') or {}).get('icon', ''),
            (e.get('eventPoint') or {}).get('lat', ''), (e.get('eventPoint') or {}).get('lng', ''),
            ';'.join(f"{p[0]},{p[1]}" for p in e.get('routPoints', []))] for i, e in enumerate(events)])
    sheet('placements', ['no', 'type', 'lat', 'lng', 'asset', 'anim', 'path'],
          [[1, 'building', '', '', 'cafe.png', 'idle', '(例) type=building/character, path=lat,lng;lat,lng']])

    os.makedirs(outdir, exist_ok=True)
    out = os.path.join(outdir, f'bus_radio_{week}.xlsx')
    wb.save(out)
    print('✅', out, '| sections:', len(rows), '| events:', len(events), '| route:', len(route))

if __name__ == '__main__':
    main(sys.argv[1], sys.argv[2])
